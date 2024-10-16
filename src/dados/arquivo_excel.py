from typing import Dict, Union, List
from src.dados.arquivo import Arquivo
from openpyxl import Workbook, load_workbook, worksheet
import os


class ArquivoExcel(Arquivo[Workbook]):
    def __init__(self, nome_aba: str):
        """init para arquivo excel

        Args:
            nome_aba (str): nome da aba da planilha
        """
        self.__planilha = Workbook()
        self.__nome_aba = nome_aba
        super().__init__()

    def __criar_cabecalho(self, dados: Dict[str, Union[str, int]], aba: worksheet) -> List[str]:
        """Métod para criar o cabeçalho das colunas

        Args:
            dados (Dict[str, Union[str, int]]): dados da req
            aba (worksheet): nome da aba

        Returns:
            List[str]: Lista de cabeçalhos
        """
        cabecalhos = list(dados.keys())
        aba.append(cabecalhos)

        return cabecalhos

    def salvar_dados(self, dados: Dict[str, Union[str, int]]):
        """Método para salvar os dados da planilha

        Args:
            dados (Dict[str, Union[str, int]]): dados 
        """
        aba = self.__planilha.active
        aba.title = self.__nome_aba

        cabecalhos = self.__criar_cabecalho(dados=dados, aba=aba)
        for linha in dados:

            valores = [linha[coluna] for coluna in cabecalhos]
            aba.append(valores)
        self.__planilha.save(self._caminho_arquivo)
        self.__planilha.close()

    def atualizar_dados(self, dados: Dict[str, Union[str, int]]):
        """Método para atualizar dados da planilha

        Args:
            dados (Dict[str, Union[str, int]]): dados da req
        """
        workbook = load_workbook(self._caminho_arquivo)
        if self.__nome_aba not in workbook.sheetnames:
            planilha = workbook.create_sheet(self.__nome_aba)
            cabecalhos = self.__criar_cabecalho(dados=dados, aba=planilha)
            for linha in dados:
                valores = [linha[coluna] for coluna in cabecalhos]
                planilha.append(valores)
        else:
            planilha = workbook[self.__nome_aba]
            ultima_lina = planilha.max_row + 1
            for _, valor in enumerate(dados, start=ultima_lina):
                planilha.append(list(valor.values()))

            ultima_lina = planilha.max_row

        workbook.save(self._caminho_arquivo)
        workbook.close()

    def verificar_arquivo(self) -> bool:
        """Método para verificar se o arquivo existe

        Returns:
            bool: verdadeiro se o arquivo existe, falso caso contrário
        """
        return os.path.exists(self._caminho_arquivo)
